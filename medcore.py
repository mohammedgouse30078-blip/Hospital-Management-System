"""
MedCore — Hospital Management System (Python / Tkinter edition)
-----------------------------------------------------------------
A desktop version of the MedCore hospital console. Every record you add —
patients, doctors, ambulance dispatches, invoices — is written straight into
a real Excel workbook (medcore_data.xlsx) that lives next to this script.
No browser, no download button: the spreadsheet updates live on disk.

Run with:  python medcore.py
Requires:  pip install openpyxl
"""

import os
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime

from openpyxl import Workbook, load_workbook

# ----------------------------------------------------------------------
# Excel persistence layer
# ----------------------------------------------------------------------

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "medcore_data.xlsx")

SHEETS = {
    "Patients": ["ID", "Name", "Age", "Gender", "Blood Group", "Ward/Room",
                 "Attending Doctor", "Contact", "Reason for Admission"],
    "Doctors": ["ID", "Name", "Specialization", "Shift", "Room/OPD", "Status"],
    "Vehicles": ["Vehicle ID"],
    "Ambulance": ["ID", "Patient/Caller", "Pickup Location", "Vehicle",
                  "Contact", "Priority", "Status"],
    "Billing": ["Invoice", "Patient", "Line Item", "Amount (INR)",
                "Invoice Total (INR)", "Status"],
}


def load_or_create_workbook():
    """Load the workbook if it exists; otherwise create it with headers."""
    if os.path.exists(DATA_FILE):
        wb = load_workbook(DATA_FILE)
        # Make sure every expected sheet/header exists (handles older files too)
        for name, headers in SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.append(headers)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > len(SHEETS):
            del wb["Sheet"]
        wb.save(DATA_FILE)
        return wb

    wb = Workbook()
    wb.remove(wb.active)
    for name, headers in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(headers)
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(header) + 2)
    wb.save(DATA_FILE)
    return wb


def append_row(sheet_name, row):
    wb = load_workbook(DATA_FILE)
    ws = wb[sheet_name]
    ws.append(row)
    wb.save(DATA_FILE)


def read_rows(sheet_name):
    wb = load_workbook(DATA_FILE)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [r for r in rows if any(v is not None for v in r)]


def overwrite_sheet(sheet_name, rows):
    """Replace all data rows in a sheet (used for delete / status updates)."""
    wb = load_workbook(DATA_FILE)
    ws = wb[sheet_name]
    ws.delete_rows(2, ws.max_row)
    for row in rows:
        ws.append(row)
    wb.save(DATA_FILE)


def next_id(sheet_name):
    rows = read_rows(sheet_name)
    ids = [r[0] for r in rows if isinstance(r[0], int)]
    return (max(ids) + 1) if ids else 1


# ----------------------------------------------------------------------
# UI theme
# ----------------------------------------------------------------------

TEAL_900 = "#0E3D3A"
TEAL_700 = "#155e58"
TEAL_100 = "#E4EFED"
BG = "#F2F4F2"
SURFACE = "#FFFFFF"
ACCENT = "#E5533C"
TEXT = "#182420"
TEXT_SOFT = "#69766F"
GOOD = "#2E8B57"
WARN = "#B8860B"


class MedCoreApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MedCore — Hospital Management System")
        self.geometry("1080x680")
        self.configure(bg=BG)
        self.minsize(920, 600)

        self._build_style()
        self._build_header()
        self._build_tabs()

        self.refresh_all()

    # ---------------- style ----------------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"),
                         padding=(16, 10), background=BG, foreground=TEXT_SOFT)
        style.map("TNotebook.Tab",
                   background=[("selected", SURFACE)],
                   foreground=[("selected", TEAL_900)])
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=26,
                         background=SURFACE, fieldbackground=SURFACE, foreground=TEXT)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"),
                         background=TEAL_100, foreground=TEAL_900)
        style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"),
                         background=TEAL_900, foreground="#fff", padding=8)
        style.map("Accent.TButton", background=[("active", TEAL_700)])
        style.configure("Danger.TButton", font=("Segoe UI", 10, "bold"),
                         background=ACCENT, foreground="#fff", padding=8)

    # ---------------- header ----------------
    def _build_header(self):
        header = tk.Frame(self, bg=TEAL_900, height=64)
        header.pack(fill="x")
        tk.Label(header, text="MedCore", font=("Segoe UI", 16, "bold"),
                  bg=TEAL_900, fg="#fff").pack(side="left", padx=20, pady=14)
        tk.Label(header, text="Hospital Operations Console", font=("Segoe UI", 10),
                  bg=TEAL_900, fg="#B9D6D1").pack(side="left")

        btn_frame = tk.Frame(header, bg=TEAL_900)
        btn_frame.pack(side="right", padx=16)
        tk.Button(btn_frame, text="Open Excel File", font=("Segoe UI", 9, "bold"),
                   bg=TEAL_700, fg="#fff", relief="flat", padx=10, pady=6,
                   activebackground=TEAL_900, activeforeground="#fff",
                   command=self.open_excel_file).pack(side="right", padx=4)
        tk.Label(btn_frame, text=f"Data file: {os.path.basename(DATA_FILE)}",
                  font=("Segoe UI", 8), bg=TEAL_900, fg="#AFD2CC").pack(side="right", padx=10)

    def open_excel_file(self):
        try:
            if sys.platform.startswith("win"):
                os.startfile(DATA_FILE)  # noqa: py-win-only
            elif sys.platform == "darwin":
                subprocess.run(["open", DATA_FILE], check=False)
            else:
                subprocess.run(["xdg-open", DATA_FILE], check=False)
        except Exception:
            messagebox.showinfo("Excel file", f"Your data is saved at:\n{DATA_FILE}")

    # ---------------- tabs ----------------
    def _build_tabs(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=14, pady=14)

        self.tab_dashboard = tk.Frame(self.notebook, bg=BG)
        self.tab_patients = tk.Frame(self.notebook, bg=BG)
        self.tab_doctors = tk.Frame(self.notebook, bg=BG)
        self.tab_ambulance = tk.Frame(self.notebook, bg=BG)
        self.tab_billing = tk.Frame(self.notebook, bg=BG)

        self.notebook.add(self.tab_dashboard, text="  Dashboard  ")
        self.notebook.add(self.tab_patients, text="  Patients  ")
        self.notebook.add(self.tab_doctors, text="  Doctor Availability  ")
        self.notebook.add(self.tab_ambulance, text="  Ambulance Service  ")
        self.notebook.add(self.tab_billing, text="  Billing  ")

        self._build_dashboard_tab()
        self._build_patients_tab()
        self._build_doctors_tab()
        self._build_ambulance_tab()
        self._build_billing_tab()

    # ==================================================================
    # DASHBOARD
    # ==================================================================
    def _build_dashboard_tab(self):
        f = self.tab_dashboard
        tk.Label(f, text="Today's operations at a glance", font=("Segoe UI", 13, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(6, 14))

        cards = tk.Frame(f, bg=BG)
        cards.pack(fill="x")
        self.stat_labels = {}
        for key, label in [("patients", "Patients on file"), ("doctors", "Doctors available"),
                            ("ambulance", "Ambulances free"), ("billing", "Invoices pending")]:
            card = tk.Frame(cards, bg=SURFACE, highlightbackground="#E5E9E6",
                              highlightthickness=1, padx=16, pady=14)
            card.pack(side="left", padx=(0, 12), fill="both", expand=True)
            tk.Label(card, text=label.upper(), font=("Segoe UI", 8, "bold"),
                      bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w")
            val = tk.Label(card, text="0", font=("Segoe UI", 22, "bold"), bg=SURFACE, fg=TEAL_900)
            val.pack(anchor="w", pady=(4, 0))
            self.stat_labels[key] = val

        tk.Label(f, text="This app writes directly into medcore_data.xlsx on every add — "
                          "no export step needed. Click 'Open Excel File' above any time.",
                  font=("Segoe UI", 9), bg=BG, fg=TEXT_SOFT, wraplength=850, justify="left"
                  ).pack(anchor="w", pady=(20, 0))

    def refresh_dashboard(self):
        patients = read_rows("Patients")
        doctors = read_rows("Doctors")
        vehicles = read_rows("Vehicles")
        dispatches = read_rows("Ambulance")
        billing = read_rows("Billing")

        available_docs = sum(1 for d in doctors if d[5] == "Available")
        busy_vehicles = {d[3] for d in dispatches if d[6] in ("En route", "On scene")}
        free_amb = max(len(vehicles) - len(busy_vehicles), 0)
        invoice_ids = {}
        for row in billing:
            invoice_ids[row[0]] = row[5]
        pending = sum(1 for status in invoice_ids.values() if status == "Pending")

        self.stat_labels["patients"].config(text=str(len(patients)))
        self.stat_labels["doctors"].config(text=f"{available_docs} / {len(doctors)}")
        self.stat_labels["ambulance"].config(text=f"{free_amb} / {len(vehicles)}")
        self.stat_labels["billing"].config(text=str(pending))

    # ==================================================================
    # PATIENTS
    # ==================================================================
    def _build_patients_tab(self):
        f = self.tab_patients
        left = tk.Frame(f, bg=SURFACE, highlightbackground="#E5E9E6", highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 14), pady=4, ipadx=10, ipady=10)
        right = tk.Frame(f, bg=BG)
        right.pack(side="left", fill="both", expand=True, pady=4)

        tk.Label(left, text="Admit a patient", font=("Segoe UI", 12, "bold"),
                  bg=SURFACE, fg=TEAL_900).pack(anchor="w", padx=14, pady=(14, 10))

        self.p_name = self._field(left, "Full name")
        self.p_age = self._field(left, "Age")
        self.p_gender = self._dropdown(left, "Gender", ["Female", "Male", "Other"])
        self.p_contact = self._field(left, "Contact number")
        self.p_blood = self._dropdown(left, "Blood group",
                                       ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"])
        self.p_ward = self._field(left, "Ward / room")
        self.p_doctor = self._dropdown(left, "Attending doctor", ["Assign later"])
        self.p_reason = self._field(left, "Reason for admission")

        ttk.Button(left, text="Admit patient", style="Accent.TButton",
                    command=self.add_patient).pack(fill="x", padx=14, pady=(12, 14))

        tk.Label(right, text="Patient records", font=("Segoe UI", 12, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(4, 8))
        cols = SHEETS["Patients"]
        self.tree_patients = ttk.Treeview(right, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree_patients.heading(c, text=c)
            self.tree_patients.column(c, width=100, anchor="w")
        self.tree_patients.pack(fill="both", expand=True)

        btn_row = tk.Frame(right, bg=BG)
        btn_row.pack(fill="x", pady=8)
        ttk.Button(btn_row, text="Remove selected", style="Danger.TButton",
                    command=self.remove_patient).pack(side="left")

    def _field(self, parent, label):
        tk.Label(parent, text=label, font=("Segoe UI", 9, "bold"), bg=SURFACE,
                  fg=TEXT_SOFT).pack(anchor="w", padx=14, pady=(6, 2))
        entry = tk.Entry(parent, font=("Segoe UI", 10), relief="solid", bd=1, width=28)
        entry.pack(padx=14, ipady=4)
        return entry

    def _dropdown(self, parent, label, values):
        tk.Label(parent, text=label, font=("Segoe UI", 9, "bold"), bg=SURFACE,
                  fg=TEXT_SOFT).pack(anchor="w", padx=14, pady=(6, 2))
        combo = ttk.Combobox(parent, values=values, state="readonly", width=26)
        combo.set(values[0])
        combo.pack(padx=14, ipady=2)
        return combo

    def add_patient(self):
        name = self.p_name.get().strip()
        age = self.p_age.get().strip()
        contact = self.p_contact.get().strip()
        if not name:
            messagebox.showwarning("Missing info", "Enter the patient's name."); return
        if not age:
            messagebox.showwarning("Missing info", "Enter the patient's age."); return
        if not contact:
            messagebox.showwarning("Missing info", "Enter a contact number."); return

        pid = next_id("Patients")
        doctor = self.p_doctor.get()
        doctor = "" if doctor == "Assign later" else doctor
        row = [pid, name, age, self.p_gender.get(), self.p_blood.get(),
               self.p_ward.get().strip(), doctor, contact, self.p_reason.get().strip()]
        append_row("Patients", row)

        for entry in (self.p_name, self.p_age, self.p_contact, self.p_ward, self.p_reason):
            entry.delete(0, tk.END)

        self.refresh_all()
        messagebox.showinfo("Admitted", f"{name} was admitted and saved to the Excel file.")

    def remove_patient(self):
        sel = self.tree_patients.selection()
        if not sel:
            messagebox.showinfo("Select a row", "Select a patient row to remove first.")
            return
        values = self.tree_patients.item(sel[0], "values")
        pid = int(values[0])
        rows = [r for r in read_rows("Patients") if r[0] != pid]
        overwrite_sheet("Patients", rows)
        self.refresh_all()

    def refresh_patients(self):
        self.tree_patients.delete(*self.tree_patients.get_children())
        for row in read_rows("Patients"):
            self.tree_patients.insert("", "end", values=row)

        doctor_names = [d[1] for d in read_rows("Doctors")]
        self.p_doctor["values"] = ["Assign later"] + doctor_names
        if self.p_doctor.get() not in self.p_doctor["values"]:
            self.p_doctor.set("Assign later")

    # ==================================================================
    # DOCTORS
    # ==================================================================
    def _build_doctors_tab(self):
        f = self.tab_doctors
        left = tk.Frame(f, bg=SURFACE, highlightbackground="#E5E9E6", highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 14), pady=4, ipadx=10, ipady=10)
        right = tk.Frame(f, bg=BG)
        right.pack(side="left", fill="both", expand=True, pady=4)

        tk.Label(left, text="Add a doctor", font=("Segoe UI", 12, "bold"),
                  bg=SURFACE, fg=TEAL_900).pack(anchor="w", padx=14, pady=(14, 10))

        self.d_name = self._field(left, "Full name")
        self.d_spec = self._field(left, "Specialization")
        self.d_status = self._dropdown(left, "Status", ["Available", "In surgery", "On leave"])
        self.d_shift = self._dropdown(left, "Shift",
                                       ["Morning (8am-2pm)", "Evening (2pm-8pm)", "Night (8pm-8am)"])
        self.d_room = self._field(left, "Room / OPD number")

        ttk.Button(left, text="Add doctor", style="Accent.TButton",
                    command=self.add_doctor).pack(fill="x", padx=14, pady=(12, 6))

        tk.Label(left, text="Update status of selected", font=("Segoe UI", 9, "bold"),
                  bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", padx=14, pady=(10, 2))
        self.d_status_update = ttk.Combobox(left, values=["Available", "In surgery", "On leave"],
                                             state="readonly", width=26)
        self.d_status_update.pack(padx=14, ipady=2)
        ttk.Button(left, text="Apply status", command=self.update_doctor_status
                    ).pack(fill="x", padx=14, pady=(8, 14))

        tk.Label(right, text="Doctor roster", font=("Segoe UI", 12, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(4, 8))
        cols = SHEETS["Doctors"]
        self.tree_doctors = ttk.Treeview(right, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree_doctors.heading(c, text=c)
            self.tree_doctors.column(c, width=110, anchor="w")
        self.tree_doctors.pack(fill="both", expand=True)

        ttk.Button(right, text="Remove selected", style="Danger.TButton",
                    command=self.remove_doctor).pack(anchor="w", pady=8)

    def add_doctor(self):
        name = self.d_name.get().strip()
        spec = self.d_spec.get().strip()
        if not name:
            messagebox.showwarning("Missing info", "Enter the doctor's name."); return
        if not spec:
            messagebox.showwarning("Missing info", "Enter a specialization."); return

        did = next_id("Doctors")
        row = [did, name, spec, self.d_status.get(), self.d_shift.get(), self.d_room.get().strip()]
        append_row("Doctors", row)

        for entry in (self.d_name, self.d_spec, self.d_room):
            entry.delete(0, tk.END)

        self.refresh_all()
        messagebox.showinfo("Added", f"Dr. {name} was added and saved to the Excel file.")

    def update_doctor_status(self):
        sel = self.tree_doctors.selection()
        new_status = self.d_status_update.get()
        if not sel or not new_status:
            messagebox.showinfo("Select a doctor", "Select a doctor row and a status first.")
            return
        values = self.tree_doctors.item(sel[0], "values")
        did = int(values[0])
        rows = read_rows("Doctors")
        updated = []
        for r in rows:
            r = list(r)
            if r[0] == did:
                r[5] = new_status
            updated.append(r)
        overwrite_sheet("Doctors", updated)
        self.refresh_all()

    def remove_doctor(self):
        sel = self.tree_doctors.selection()
        if not sel:
            messagebox.showinfo("Select a row", "Select a doctor row to remove first.")
            return
        values = self.tree_doctors.item(sel[0], "values")
        did = int(values[0])
        rows = [r for r in read_rows("Doctors") if r[0] != did]
        overwrite_sheet("Doctors", rows)
        self.refresh_all()

    def refresh_doctors(self):
        self.tree_doctors.delete(*self.tree_doctors.get_children())
        for row in read_rows("Doctors"):
            self.tree_doctors.insert("", "end", values=row)

    # ==================================================================
    # AMBULANCE
    # ==================================================================
    def _build_ambulance_tab(self):
        f = self.tab_ambulance
        left = tk.Frame(f, bg=SURFACE, highlightbackground="#E5E9E6", highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 14), pady=4, ipadx=10, ipady=10)
        right = tk.Frame(f, bg=BG)
        right.pack(side="left", fill="both", expand=True, pady=4)

        tk.Label(left, text="Dispatch an ambulance", font=("Segoe UI", 12, "bold"),
                  bg=SURFACE, fg=TEAL_900).pack(anchor="w", padx=14, pady=(14, 10))

        self.a_name = self._field(left, "Patient / caller name")
        self.a_pickup = self._field(left, "Pickup location")
        self.a_vehicle = self._dropdown(left, "Assign vehicle", ["—"])
        self.a_contact = self._field(left, "Contact number")
        self.a_priority = self._dropdown(left, "Priority",
                                          ["Emergency", "Urgent", "Routine transfer"])

        ttk.Button(left, text="Dispatch", style="Danger.TButton",
                    command=self.add_dispatch).pack(fill="x", padx=14, pady=(12, 6))

        tk.Label(left, text="Update status of selected", font=("Segoe UI", 9, "bold"),
                  bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", padx=14, pady=(10, 2))
        self.a_status_update = ttk.Combobox(left, values=["En route", "On scene", "Completed"],
                                             state="readonly", width=26)
        self.a_status_update.pack(padx=14, ipady=2)
        ttk.Button(left, text="Apply status", command=self.update_dispatch_status
                    ).pack(fill="x", padx=14, pady=(8, 14))

        tk.Label(left, text="Add vehicle to fleet", font=("Segoe UI", 9, "bold"),
                  bg=SURFACE, fg=TEXT_SOFT).pack(anchor="w", padx=14, pady=(10, 2))
        veh_row = tk.Frame(left, bg=SURFACE)
        veh_row.pack(padx=14, pady=(0, 14), fill="x")
        self.new_vehicle = tk.Entry(veh_row, font=("Segoe UI", 10), relief="solid", bd=1, width=18)
        self.new_vehicle.pack(side="left", ipady=4)
        ttk.Button(veh_row, text="Add", command=self.add_vehicle).pack(side="left", padx=6)

        tk.Label(right, text="Dispatch log", font=("Segoe UI", 12, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(4, 8))
        cols = SHEETS["Ambulance"]
        self.tree_ambulance = ttk.Treeview(right, columns=cols, show="headings", height=14)
        for c in cols:
            self.tree_ambulance.heading(c, text=c)
            self.tree_ambulance.column(c, width=110, anchor="w")
        self.tree_ambulance.pack(fill="both", expand=True)

        ttk.Button(right, text="Remove selected", style="Danger.TButton",
                    command=self.remove_dispatch).pack(anchor="w", pady=8)

        tk.Label(right, text="Fleet", font=("Segoe UI", 11, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(6, 4))
        self.tree_fleet = ttk.Treeview(right, columns=("Vehicle ID", "Status"),
                                        show="headings", height=4)
        self.tree_fleet.heading("Vehicle ID", text="Vehicle ID")
        self.tree_fleet.heading("Status", text="Status")
        self.tree_fleet.pack(fill="x")

    def add_vehicle(self):
        vid = self.new_vehicle.get().strip()
        if not vid:
            return
        existing = {r[0] for r in read_rows("Vehicles")}
        if vid in existing:
            messagebox.showinfo("Already exists", "That vehicle ID is already in the fleet.")
            return
        append_row("Vehicles", [vid])
        self.new_vehicle.delete(0, tk.END)
        self.refresh_all()

    def add_dispatch(self):
        name = self.a_name.get().strip()
        pickup = self.a_pickup.get().strip()
        contact = self.a_contact.get().strip()
        vehicle = self.a_vehicle.get()
        if not name:
            messagebox.showwarning("Missing info", "Enter a patient or caller name."); return
        if not pickup:
            messagebox.showwarning("Missing info", "Enter the pickup location."); return
        if not contact:
            messagebox.showwarning("Missing info", "Enter a contact number."); return
        if vehicle == "—":
            messagebox.showwarning("Missing info", "Add a vehicle to the fleet first."); return

        did = next_id("Ambulance")
        row = [did, name, pickup, vehicle, contact, self.a_priority.get(), "En route"]
        append_row("Ambulance", row)

        for entry in (self.a_name, self.a_pickup, self.a_contact):
            entry.delete(0, tk.END)

        self.refresh_all()
        messagebox.showinfo("Dispatched", "Ambulance dispatched and logged to the Excel file.")

    def update_dispatch_status(self):
        sel = self.tree_ambulance.selection()
        new_status = self.a_status_update.get()
        if not sel or not new_status:
            messagebox.showinfo("Select a dispatch", "Select a dispatch row and a status first.")
            return
        values = self.tree_ambulance.item(sel[0], "values")
        did = int(values[0])
        rows = read_rows("Ambulance")
        updated = []
        for r in rows:
            r = list(r)
            if r[0] == did:
                r[6] = new_status
            updated.append(r)
        overwrite_sheet("Ambulance", updated)
        self.refresh_all()

    def remove_dispatch(self):
        sel = self.tree_ambulance.selection()
        if not sel:
            messagebox.showinfo("Select a row", "Select a dispatch row to remove first.")
            return
        values = self.tree_ambulance.item(sel[0], "values")
        did = int(values[0])
        rows = [r for r in read_rows("Ambulance") if r[0] != did]
        overwrite_sheet("Ambulance", rows)
        self.refresh_all()

    def refresh_ambulance(self):
        self.tree_ambulance.delete(*self.tree_ambulance.get_children())
        for row in read_rows("Ambulance"):
            self.tree_ambulance.insert("", "end", values=row)

        vehicles = [r[0] for r in read_rows("Vehicles")]
        self.a_vehicle["values"] = vehicles if vehicles else ["—"]
        if self.a_vehicle.get() not in self.a_vehicle["values"]:
            self.a_vehicle.set(self.a_vehicle["values"][0])

        busy = {r[3] for r in read_rows("Ambulance") if r[6] in ("En route", "On scene")}
        self.tree_fleet.delete(*self.tree_fleet.get_children())
        for vid in vehicles:
            status = "On call" if vid in busy else "Available"
            self.tree_fleet.insert("", "end", values=(vid, status))

    # ==================================================================
    # BILLING
    # ==================================================================
    def _build_billing_tab(self):
        f = self.tab_billing
        left = tk.Frame(f, bg=SURFACE, highlightbackground="#E5E9E6", highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 14), pady=4, ipadx=10, ipady=10)
        right = tk.Frame(f, bg=BG)
        right.pack(side="left", fill="both", expand=True, pady=4)

        tk.Label(left, text="Create invoice", font=("Segoe UI", 12, "bold"),
                  bg=SURFACE, fg=TEAL_900).pack(anchor="w", padx=14, pady=(14, 10))

        self.b_patient = self._dropdown(left, "Patient", ["—"])

        item_row = tk.Frame(left, bg=SURFACE)
        item_row.pack(padx=14, pady=(6, 0), fill="x")
        tk.Label(item_row, text="Line item", font=("Segoe UI", 9, "bold"),
                  bg=SURFACE, fg=TEXT_SOFT).grid(row=0, column=0, sticky="w")
        tk.Label(item_row, text="Amount", font=("Segoe UI", 9, "bold"),
                  bg=SURFACE, fg=TEXT_SOFT).grid(row=0, column=1, sticky="w", padx=(8, 0))
        self.b_item_name = tk.Entry(item_row, font=("Segoe UI", 10), relief="solid", bd=1, width=17)
        self.b_item_name.grid(row=1, column=0, ipady=4)
        self.b_item_amount = tk.Entry(item_row, font=("Segoe UI", 10), relief="solid", bd=1, width=9)
        self.b_item_amount.grid(row=1, column=1, ipady=4, padx=(8, 0))
        ttk.Button(left, text="+ Add line item", command=self.add_bill_item
                    ).pack(fill="x", padx=14, pady=(8, 8))

        self.bill_items_list = tk.Listbox(left, height=6, font=("Segoe UI", 9), width=32)
        self.bill_items_list.pack(padx=14, fill="x")
        self.bill_total_label = tk.Label(left, text="Total: ₹0", font=("Segoe UI", 12, "bold"),
                                          bg=SURFACE, fg=TEAL_900)
        self.bill_total_label.pack(anchor="w", padx=14, pady=(8, 4))

        self._current_bill_items = []

        ttk.Button(left, text="Generate invoice", style="Accent.TButton",
                    command=self.generate_invoice).pack(fill="x", padx=14, pady=(6, 14))

        tk.Label(right, text="Invoices", font=("Segoe UI", 12, "bold"),
                  bg=BG, fg=TEAL_900).pack(anchor="w", pady=(4, 8))
        cols = ("Invoice", "Patient", "Total (INR)", "Status")
        self.tree_billing = ttk.Treeview(right, columns=cols, show="headings", height=16)
        for c in cols:
            self.tree_billing.heading(c, text=c)
            self.tree_billing.column(c, width=140, anchor="w")
        self.tree_billing.pack(fill="both", expand=True)

        btn_row = tk.Frame(right, bg=BG)
        btn_row.pack(fill="x", pady=8)
        ttk.Button(btn_row, text="Mark selected as Paid", style="Accent.TButton",
                    command=self.mark_invoice_paid).pack(side="left")

    def add_bill_item(self):
        name = self.b_item_name.get().strip()
        amount = self.b_item_amount.get().strip()
        if not name or not amount:
            messagebox.showwarning("Missing info", "Enter both an item name and an amount.")
            return
        try:
            amount = float(amount)
        except ValueError:
            messagebox.showwarning("Invalid amount", "Amount must be a number.")
            return
        self._current_bill_items.append((name, amount))
        self.bill_items_list.insert(tk.END, f"{name} — ₹{amount:,.0f}")
        self.b_item_name.delete(0, tk.END)
        self.b_item_amount.delete(0, tk.END)
        total = sum(a for _, a in self._current_bill_items)
        self.bill_total_label.config(text=f"Total: ₹{total:,.0f}")

    def generate_invoice(self):
        patient = self.b_patient.get()
        if patient == "—":
            messagebox.showwarning("Missing info", "Select a patient first."); return
        if not self._current_bill_items:
            messagebox.showwarning("Missing info", "Add at least one line item first."); return

        rows = read_rows("Billing")
        existing_invoice_nums = set()
        for r in rows:
            if isinstance(r[0], str) and r[0].startswith("INV-"):
                existing_invoice_nums.add(int(r[0].split("-")[1]))
        invoice_num = (max(existing_invoice_nums) + 1) if existing_invoice_nums else 1
        invoice_id = f"INV-{invoice_num:04d}"

        total = sum(a for _, a in self._current_bill_items)
        for name, amount in self._current_bill_items:
            append_row("Billing", [invoice_id, patient, name, amount, total, "Pending"])

        self._current_bill_items = []
        self.bill_items_list.delete(0, tk.END)
        self.bill_total_label.config(text="Total: ₹0")

        self.refresh_all()
        messagebox.showinfo("Invoice generated", f"{invoice_id} saved to the Excel file.")

    def mark_invoice_paid(self):
        sel = self.tree_billing.selection()
        if not sel:
            messagebox.showinfo("Select a row", "Select an invoice row first.")
            return
        invoice_id = self.tree_billing.item(sel[0], "values")[0]
        rows = read_rows("Billing")
        updated = []
        for r in rows:
            r = list(r)
            if r[0] == invoice_id:
                r[5] = "Paid"
            updated.append(r)
        overwrite_sheet("Billing", updated)
        self.refresh_all()

    def refresh_billing(self):
        self.tree_billing.delete(*self.tree_billing.get_children())
        rows = read_rows("Billing")
        seen = {}
        for r in rows:
            invoice_id, patient, _item, _amount, total, status = r
            seen[invoice_id] = (patient, total, status)
        for invoice_id, (patient, total, status) in seen.items():
            self.tree_billing.insert("", "end", values=(invoice_id, patient, f"{total:,.0f}", status))

        patient_names = [p[1] for p in read_rows("Patients")]
        self.b_patient["values"] = ["—"] + patient_names
        if self.b_patient.get() not in self.b_patient["values"]:
            self.b_patient.set("—")

    # ==================================================================
    def refresh_all(self):
        self.refresh_patients()
        self.refresh_doctors()
        self.refresh_ambulance()
        self.refresh_billing()
        self.refresh_dashboard()


if __name__ == "__main__":
    load_or_create_workbook()
    app = MedCoreApp()
    app.mainloop()
